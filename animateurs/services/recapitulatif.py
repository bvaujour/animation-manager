"""Synthèse des jours planifiés et de la paie par animateur et par lieu."""

from __future__ import annotations

import datetime
from collections import defaultdict
from decimal import Decimal
from io import BytesIO

from django.utils import timezone

from animateurs.models import Affectation, PeriodeScolaire
from animateurs.services.flottants import est_groupe_flottants
from animateurs.services.temps_travail import (
    activites_temps_travail_pour_periodes,
    comptabiliser_jours_complementaires,
)


def _jours_entre(debut: datetime.date, fin_exclusive: datetime.date):
    """Itère de ``debut`` inclus à ``fin_exclusive`` exclu."""

    jour = debut
    while jour < fin_exclusive:
        yield jour
        jour += datetime.timedelta(days=1)


def _montant(jours: int, paie_jour):
    """Calcule un montant sérialisable, ou ``None`` si le tarif manque."""

    if paie_jour is None:
        return None
    return str((Decimal(jours) * paie_jour).quantize(Decimal("0.01")))


def _nombre_json(nombre):
    """Conserve des entiers lisibles tout en autorisant les demi-journées."""

    valeur = Decimal(nombre)
    return int(valeur) if valeur == valeur.to_integral_value() else float(valeur)


def generer_recapitulatif(debut, fin, jours_selectionnes=None, periode_ids=None):
    """Retourne les jours et la paie par animateur, ventilés par centre.

    ``debut`` est inclus et ``fin`` est exclusif. Une même date ne compte
    qu'une fois par animateur et par centre, même si plusieurs groupes du même
    centre lui sont affectés ce jour-là. Le total animateur compte également
    chaque date une seule fois.
    """

    debut_date = debut.date()
    fin_date = fin.date()
    jours_autorises = set(_jours_entre(debut_date, fin_date))
    if jours_selectionnes is not None:
        jours_autorises &= set(jours_selectionnes)

    affectations = (
        Affectation.objects.select_related("animateur", "centre", "evenement")
        .filter(debut__lt=fin, fin__gt=debut)
        .order_by("animateur__prenom", "animateur__nom", "debut")
    )

    jours_par_animateur = defaultdict(set)
    jours_par_animateur_centre = defaultdict(lambda: defaultdict(set))
    details_par_animateur = defaultdict(lambda: defaultdict(dict))
    animateurs = {}
    centres = {}

    for affectation in affectations:
        animateur = affectation.animateur
        centre = affectation.centre
        animateurs[animateur.id] = animateur
        centres[centre.id] = {
            "id": centre.id,
            "nom": centre.nom,
            "code": centre.code,
            "couleur": centre.couleur,
            "ordre": centre.ordre,
        }

        debut_affectation = max(timezone.localtime(affectation.debut).date(), debut_date)
        fin_affectation = min(timezone.localtime(affectation.fin).date(), fin_date)
        for jour in _jours_entre(debut_affectation, fin_affectation):
            if jour not in jours_autorises:
                continue
            jours_par_animateur[animateur.id].add(jour)
            jours_par_animateur_centre[animateur.id][centre.id].add(jour)
            details_par_animateur[animateur.id][jour][centre.id] = {
                "id": centre.id,
                "nom": centre.nom,
                "code": centre.code,
                "couleur": centre.couleur,
                "groupe": "Animateur flottant" if est_groupe_flottants(affectation.evenement) else affectation.evenement.nom,
            }

    centres_tries = sorted(
        centres.values(),
        key=lambda centre: (centre["ordre"], centre["nom"].casefold(), centre["code"]),
    )

    ids_periodes_selectionnees = {int(item) for item in (periode_ids or [])}
    if ids_periodes_selectionnees:
        periodes_travail = list(PeriodeScolaire.objects.filter(pk__in=ids_periodes_selectionnees))
    else:
        periodes_travail = list(PeriodeScolaire.objects.filter(
            debut__gte=debut_date, fin__lt=fin_date
        ))
    activites = activites_temps_travail_pour_periodes(
        periodes_travail,
        accepter_selection_englobante=not ids_periodes_selectionnees,
    )
    complements = comptabiliser_jours_complementaires(
        animateurs.keys(), jours_par_animateur, activites
    )

    lignes = []
    for animateur in animateurs.values():
        jours_totaux = jours_par_animateur[animateur.id]
        if not jours_totaux:
            continue
        ventilation = []
        for centre in centres_tries:
            nombre_jours = len(jours_par_animateur_centre[animateur.id][centre["id"]])
            ventilation.append({
                "centre_id": centre["id"],
                "jours_travailles": nombre_jours,
                "paie": _montant(nombre_jours, animateur.paie_jour),
            })

        complement = complements[animateur.id]
        jours_reunion = complement["reunion"]
        jours_preparation = complement["preparation"]
        dates_reunions_comptabilisees = [
            jour.isoformat() for jour in sorted(complement["dates_reunions"])
        ]
        jours_affectation = Decimal(len(jours_totaux))
        total_jours = jours_affectation + jours_reunion + jours_preparation
        lignes.append({
            "id": animateur.id,
            "prenom": animateur.prenom,
            "nom": animateur.nom,
            "paie_jour": str(animateur.paie_jour) if animateur.paie_jour is not None else None,
            "jours_affectation": _nombre_json(jours_affectation),
            "jours_reunion": _nombre_json(jours_reunion),
            "jours_preparation": _nombre_json(jours_preparation),
            "dates_reunions_comptabilisees": dates_reunions_comptabilisees,
            "jours_travailles": _nombre_json(total_jours),
            "paie_totale": _montant(total_jours, animateur.paie_jour),
            "centres": ventilation,
            "jours": [
                {
                    "date": jour.isoformat(),
                    "lieux": list(details_par_animateur[animateur.id][jour].values()),
                }
                for jour in sorted(jours_totaux)
            ],
        })

    lignes.sort(key=lambda ligne: (ligne["prenom"].casefold(), ligne["nom"].casefold()))
    total_paie = sum(
        (Decimal(ligne["paie_totale"]) for ligne in lignes if ligne["paie_totale"] is not None),
        Decimal("0.00"),
    )

    return {
        "dates": [jour.isoformat() for jour in sorted(jours_autorises)],
        "centres": [{key: value for key, value in centre.items() if key != "ordre"} for centre in centres_tries],
        "animateurs": lignes,
        "total_jours": _nombre_json(sum((Decimal(str(ligne["jours_travailles"])) for ligne in lignes), Decimal("0"))),
        "total_paie_connue": str(total_paie.quantize(Decimal("0.01"))),
        "tarifs_manquants": sum(1 for ligne in lignes if ligne["paie_jour"] is None),
    }


def _euros_pdf(valeur):
    if valeur is None:
        return "Manquant"
    return f"{Decimal(valeur):,.2f} €".replace(",", " ").replace(".", ",")


ENTETES_PREPARATION_PAIE = [
    "Animateur", "Affectations", "Réunions", "Télétravail / préparation",
    "Total jours", "Base CEE", "CP CEE", "Salaire mensuel de référence",
    "Primes", "Total préparé",
]


def _total_prepare_affiche(animateur):
    if animateur.get("total_prepare") is not None:
        if (
            animateur.get("paie_habituelle")
            and Decimal(animateur.get("total_prepare", "0")) == 0
        ):
            return "Paie habituelle"
        return _euros_pdf(animateur["total_prepare"])
    return "Incomplet" if animateur.get("etat_preparation") == "incomplet" else "—"


def lignes_recapitulatif_paie(recap):
    """Construit le même tableau agrégé que l'onglet Totaux par animateur."""
    lignes = [ENTETES_PREPARATION_PAIE]
    for animateur in recap["animateurs"]:
        lignes.append([
            f'{animateur["prenom"]} {animateur["nom"]}',
            str(animateur["jours_affectation"]), str(animateur["jours_reunion"]),
            str(animateur["jours_preparation"]), str(animateur["jours_travailles"]),
            _euros_pdf(animateur.get("base_cee", 0)),
            _euros_pdf(animateur.get("indemnite_cp_cee", 0)),
            _euros_pdf(animateur.get("salaire_mensuel_reference"))
            if animateur.get("salaire_mensuel_reference") is not None else "—",
            _euros_pdf(animateur.get("montant_primes_preparees", 0)),
            _total_prepare_affiche(animateur),
        ])
    references_mensuelles = [
        item for item in recap["animateurs"] if item.get("salaire_mensuel_reference") is not None
    ]
    total_mensuel = (
        sum((Decimal(item["base_mensuelle_reference"]) for item in references_mensuelles), Decimal("0"))
        if references_mensuelles
        and all(item.get("base_mensuelle_reference") is not None for item in references_mensuelles)
        else None
    )
    lignes.append([
        "TOTAL",
        str(sum(Decimal(str(item["jours_affectation"])) for item in recap["animateurs"])),
        str(sum(Decimal(str(item["jours_reunion"])) for item in recap["animateurs"])),
        str(sum(Decimal(str(item["jours_preparation"])) for item in recap["animateurs"])),
        str(recap["total_jours"]),
        _euros_pdf(sum((Decimal(item.get("base_cee", "0")) for item in recap["animateurs"]), Decimal("0"))),
        _euros_pdf(sum((Decimal(item.get("indemnite_cp_cee", "0")) for item in recap["animateurs"]), Decimal("0"))),
        _euros_pdf(total_mensuel) if total_mensuel is not None else "—",
        _euros_pdf(recap.get("total_primes_preparees", 0)),
        _euros_pdf(recap.get("total_prepare", 0)),
    ])
    return lignes


def alertes_recapitulatif_paie(recap):
    """Expose uniquement les alertes utiles, sans répéter les montants du tableau."""
    return [
        {
            "animateur": f'{item["prenom"]} {item["nom"]}',
            "messages": [alerte["message"] for alerte in item.get("alertes_paie", [])],
        }
        for item in recap["animateurs"] if item.get("alertes_paie")
    ]


def generer_recapitulatif_paie_pdf(recap, debut: datetime.date, fin: datetime.date) -> bytes:
    """Crée un PDF compact reprenant les totaux de paie de la sélection."""

    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    output = BytesIO()
    document = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        leftMargin=14 * mm,
        rightMargin=14 * mm,
        topMargin=14 * mm,
        bottomMargin=14 * mm,
        title=f"Récapitulatif de paie du {debut:%d/%m/%Y} au {fin:%d/%m/%Y}",
    )
    styles = getSampleStyleSheet()
    titre = ParagraphStyle(
        "RecapPaieTitle",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=16,
        textColor=colors.HexColor("#1F6F54"),
        alignment=TA_CENTER,
        spaceAfter=5 * mm,
    )

    lignes = lignes_recapitulatif_paie(recap)
    lignes[0] = [Paragraph(item, styles["Normal"]) for item in lignes[0]]

    tableau = Table(lignes, colWidths=[35 * mm, 15 * mm, 15 * mm, 28 * mm, 16 * mm, 21 * mm, 18 * mm, 31 * mm, 20 * mm, 27 * mm], repeatRows=1)
    tableau.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F6F54")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#E8F3EE")),
        ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D7E2DC")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#F7FAF8")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
        ("FONTSIZE", (0, 0), (-1, -1), 7.5),
    ]))

    contenu = [
        Paragraph("Récapitulatif de paie", titre),
        Paragraph(f"Période du {debut:%d/%m/%Y} au {fin:%d/%m/%Y}", styles["Heading2"]),
        Spacer(1, 4 * mm),
    ]
    if recap["animateurs"]:
        contenu.append(tableau)
        animateurs_alertes = alertes_recapitulatif_paie(recap)
        if animateurs_alertes:
            contenu.extend([Spacer(1, 4 * mm), Paragraph("Alertes / éléments à vérifier", styles["Heading2"])])
            for animateur in animateurs_alertes:
                alertes = " ; ".join(animateur["messages"])
                contenu.append(Paragraph(
                    f'<b>{animateur["animateur"]}</b> — {alertes}',
                    styles["Normal"],
                ))
    else:
        contenu.append(Paragraph("Aucune journée planifiée sur cette période.", styles["Normal"]))
    document.build(contenu)
    output.seek(0)
    return output.read()


def generer_recapitulatif_excel(recap, debut: datetime.date, fin: datetime.date) -> bytes:
    """Crée un classeur exploitable reprenant les totaux et la ventilation par centre."""

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    classeur = Workbook()
    feuille = classeur.active
    feuille.title = "Totaux paie"
    entetes = ENTETES_PREPARATION_PAIE
    feuille.append([f"Récapitulatif du {debut:%d/%m/%Y} au {fin:%d/%m/%Y}"])
    feuille.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(entetes))
    feuille.append(entetes)
    for animateur in recap["animateurs"]:
        feuille.append([
            f'{animateur["prenom"]} {animateur["nom"]}',
            animateur["jours_affectation"], animateur["jours_reunion"],
            animateur["jours_preparation"], animateur["jours_travailles"],
            float(Decimal(animateur.get("base_cee", "0"))),
            float(Decimal(animateur.get("indemnite_cp_cee", "0"))),
            float(Decimal(animateur["salaire_mensuel_reference"]))
            if animateur.get("salaire_mensuel_reference") is not None else None,
            float(Decimal(animateur.get("montant_primes_preparees", "0"))),
            float(Decimal(animateur["total_prepare"]))
            if animateur.get("total_prepare") is not None else (
                "Paie habituelle" if animateur.get("paie_habituelle") else "Incomplet"
            ),
        ])
    total_export = lignes_recapitulatif_paie(recap)[-1]
    feuille.append([total_export[0], *[
        float(Decimal(str(value).replace(" €", "").replace(" ", "").replace(",", ".")))
        if index > 0 and value != "—" else None if value == "—" else value
        for index, value in enumerate(total_export[1:], start=1)
    ]])

    centres = classeur.create_sheet("Détail par centre")
    centres.append(["Animateur", *[centre["nom"] for centre in recap["centres"]], "Total jours"])
    for animateur in recap["animateurs"]:
        jours_par_centre = {item["centre_id"]: item["jours_travailles"] for item in animateur["centres"]}
        centres.append([
            f'{animateur["prenom"]} {animateur["nom"]}',
            *[jours_par_centre.get(centre["id"], 0) for centre in recap["centres"]],
            animateur["jours_travailles"],
        ])

    preparation = classeur.create_sheet("Préparation contrats")
    preparation.append([
        "Animateur", "Contrat", "Base CEE", "Indemnité CP CEE",
        "Salaire mensuel de référence", "Primes attribuées", "Total préparé",
        "État", "Alertes / informations",
    ])
    for animateur in recap["animateurs"]:
        preparation.append([
            f'{animateur["prenom"]} {animateur["nom"]}',
            animateur.get("type_contrat_libelle", ""),
            float(Decimal(animateur.get("base_cee", "0"))),
            float(Decimal(animateur.get("indemnite_cp_cee", "0"))),
            float(Decimal(animateur["salaire_mensuel_reference"]))
            if animateur.get("salaire_mensuel_reference") is not None else None,
            float(Decimal(animateur.get("montant_primes_preparees", "0"))),
            float(Decimal(animateur["total_prepare"])) if animateur.get("total_prepare") is not None else None,
            animateur.get("etat_preparation", ""),
            " | ".join(item["message"] for item in animateur.get("alertes_paie", [])),
        ])

    vert = "1F6F54"
    for onglet, ligne_entete in ((feuille, 2), (centres, 1), (preparation, 1)):
        onglet.freeze_panes = f"A{ligne_entete + 1}"
        onglet.auto_filter.ref = f"A{ligne_entete}:{get_column_letter(onglet.max_column)}{onglet.max_row}"
        for cellule in onglet[ligne_entete]:
            cellule.fill = PatternFill("solid", fgColor=vert)
            cellule.font = Font(color="FFFFFF", bold=True)
            cellule.alignment = Alignment(horizontal="center", vertical="center")
        for colonne in range(1, onglet.max_column + 1):
            valeurs = [str(onglet.cell(ligne, colonne).value or "") for ligne in range(1, onglet.max_row + 1)]
            onglet.column_dimensions[get_column_letter(colonne)].width = min(34, max(12, max(map(len, valeurs)) + 2))
    feuille["A1"].font = Font(size=15, bold=True, color=vert)
    feuille["A1"].alignment = Alignment(horizontal="center")
    for ligne in range(3, feuille.max_row + 1):
        for colonne in range(6, 11):
            feuille.cell(ligne, colonne).number_format = '#,##0.00 [$€-fr-FR]'
    for cellule in feuille[feuille.max_row]:
        cellule.font = Font(bold=True)

    sortie = BytesIO()
    classeur.save(sortie)
    return sortie.getvalue()
