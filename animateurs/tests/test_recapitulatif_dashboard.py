import datetime
from decimal import Decimal
from io import BytesIO, StringIO
from pathlib import Path

from django.conf import settings
from django.core.management import call_command
from django.urls import reverse
from django.utils import timezone

from animateurs.models import (
    Affectation, ActiviteTravailComplementaire, Animateur, Centre, Evenement,
    ParticipationTravailComplementaire, PrimeJournalierePeriode,
)
from animateurs.tests.base import ConnexionTestCase
from animateurs.tests.factories import creer_periode
from animateurs.services.recapitulatif import (
    alertes_recapitulatif_paie,
    generer_recapitulatif_paie_pdf,
    lignes_recapitulatif_paie,
)


class RecapitulatifDashboardTests(ConnexionTestCase):
    def setUp(self):
        self.centre = Centre.objects.create(nom="La Pacaudière", code="PAC", couleur="#123456")
        self.periode = creer_periode(debut=datetime.date(2026, 7, 6), nom="Semaine récap")
        self.evenement = Evenement.objects.create(
            centre=self.centre,
            nom="Maternelles",
            effectif_cible=2,
            jours_ouverts=[0, 1, 2, 3, 4],
        )
        self.evenement.periodes_scolaires.add(self.periode)
        self.julie = Animateur.objects.create(prenom="Julie", nom="Martin", paie_jour="65.00")
        self.sam = Animateur.objects.create(prenom="Sam", nom="Dupont")

    def _affecter(self, animateur, jour, duree=1):
        debut = timezone.make_aware(datetime.datetime.combine(jour, datetime.time.min))
        return Affectation.objects.create(
            animateur=animateur,
            centre=self.centre,
            evenement=self.evenement,
            debut=debut,
            fin=debut + datetime.timedelta(days=duree),
        )

    def test_api_compte_les_jours_travailles_par_animateur(self):
        self._affecter(self.julie, datetime.date(2026, 7, 6), duree=2)
        self._affecter(self.sam, datetime.date(2026, 7, 7))

        response = self.client.get(reverse("api_recapitulatif") + "?debut=2026-07-06&fin=2026-07-09")

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data["total_jours"], 3)
        self.assertNotIn("synthese", data)
        self.assertNotIn("alertes", data)
        self.assertNotIn("evenements", data)

        julie = next(item for item in data["animateurs"] if item["id"] == self.julie.id)
        sam = next(item for item in data["animateurs"] if item["id"] == self.sam.id)
        self.assertEqual(julie["jours_travailles"], 2)
        self.assertEqual(sam["jours_travailles"], 1)

    def test_une_date_ne_compte_quune_fois_par_animateur(self):
        self._affecter(self.julie, datetime.date(2026, 7, 6))
        autre_groupe = Evenement.objects.create(
            centre=self.centre,
            nom="Élémentaires",
            effectif_cible=1,
        )
        debut = timezone.make_aware(datetime.datetime(2026, 7, 6))
        Affectation.objects.create(
            animateur=self.julie,
            centre=self.centre,
            evenement=autre_groupe,
            debut=debut,
            fin=debut + datetime.timedelta(days=1),
        )

        data = self.client.get(reverse("api_recapitulatif") + "?debut=2026-07-06&fin=2026-07-07").json()
        self.assertEqual(data["animateurs"][0]["jours_travailles"], 1)
        self.assertEqual(data["total_jours"], 1)

    def test_api_accepte_plusieurs_periodes_discontinues(self):
        seconde_periode = creer_periode(debut=datetime.date(2026, 7, 20), nom="Deuxième semaine récap")
        self._affecter(self.julie, datetime.date(2026, 7, 6))
        self._affecter(self.julie, datetime.date(2026, 7, 20))
        self._affecter(self.julie, datetime.date(2026, 7, 13))

        response = self.client.get(
            reverse("api_recapitulatif") + f"?periode_ids={self.periode.id},{seconde_periode.id}"
        )

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data["periode"]["ids"], [self.periode.id, seconde_periode.id])
        self.assertEqual(data["animateurs"][0]["jours_travailles"], 2)


    def test_api_retourne_le_detail_des_lieux_par_jour(self):
        self._affecter(self.julie, datetime.date(2026, 7, 6), duree=2)

        data = self.client.get(
            reverse("api_recapitulatif") + f"?periode_ids={self.periode.id}"
        ).json()

        self.assertEqual(data["dates"][0], "2026-07-06")
        self.assertEqual(data["centres"], [{
            "id": self.centre.id,
            "nom": "La Pacaudière",
            "code": "PAC",
            "couleur": "#123456",
        }])
        julie = next(item for item in data["animateurs"] if item["id"] == self.julie.id)
        self.assertEqual([jour["date"] for jour in julie["jours"]], ["2026-07-06", "2026-07-07"])
        self.assertEqual(julie["jours"][0]["lieux"][0]["code"], "PAC")
        self.assertEqual(julie["jours"][0]["lieux"][0]["couleur"], "#123456")
        self.assertEqual(julie["centres"][0]["jours_travailles"], 2)
        self.assertEqual(julie["centres"][0]["paie"], "130.00")
        self.assertEqual(julie["paie_totale"], "130.00")

    def test_page_paie_affiche_les_quatre_onglets_dans_le_bon_ordre(self):
        response = self.client.get(reverse("recapitulatif"))
        contenu = response.content.decode()

        self.assertContains(response, "Paie")
        self.assertContains(response, "Temps de travail")
        self.assertContains(response, "Prime")
        self.assertContains(response, "Jours et paie par centre")
        self.assertContains(response, "Totaux par animateur")
        self.assertLess(contenu.index("Temps de travail"), contenu.index("Prime"))
        self.assertLess(contenu.index("Prime"), contenu.index("Jours et paie par centre"))
        self.assertLess(contenu.index("Jours et paie par centre"), contenu.index("Totaux par animateur"))
        self.assertContains(response, 'data-recap-panel="temps-travail"')
        self.assertContains(response, 'data-recap-panel="centres"')
        self.assertNotContains(response, 'id="recap-period-selector"')
        self.assertContains(response, 'id="recap-date-start"')
        self.assertContains(response, 'id="recap-date-end"')
        self.assertContains(response, 'id="recap-refresh"')
        self.assertContains(response, 'data-week-picker-mode="multiple"')
        self.assertNotContains(response, 'id="app-type-accueil"')
        self.assertNotContains(response, 'id="app-periode-accueil"')
        self.assertContains(response, 'id="btn-recap-pdf"')
        self.assertContains(response, 'id="btn-recap-excel"')
        self.assertContains(response, 'data-recap-panel="paie"')
        self.assertNotContains(response, 'data-recap-panel="temps-travail" hidden')
        self.assertContains(response, 'data-recap-panel="centres" hidden')
        self.assertContains(response, 'data-recap-panel="totaux" hidden')
        self.assertContains(response, 'data-recap-panel="paie" hidden')

    def test_interface_initialise_les_bornes_au_mois_civil_courant(self):
        javascript = (Path(settings.BASE_DIR) / "static/js/recapitulatif.js").read_text(encoding="utf-8")

        self.assertIn("new Date(today.getFullYear(), today.getMonth(), 1)", javascript)
        self.assertIn("new Date(today.getFullYear(), today.getMonth() + 1, 0)", javascript)
        self.assertIn('refreshButton?.addEventListener("click"', javascript)

    def test_entete_recapitulatif_reste_sur_une_ligne_en_affichage_ordinateur(self):
        response = self.client.get(reverse("recapitulatif"))
        contenu = response.content.decode()
        css = (Path(settings.BASE_DIR) / "static/css/recapitulatif.css").read_text(encoding="utf-8")

        titre = contenu.index('class="app-page-title"')
        filtres = contenu.index('class="recap-date-filters"')
        exports = contenu.index("recap-exports")
        self.assertLess(titre, filtres)
        self.assertLess(filtres, exports)
        self.assertIn("@media (min-width:800px){.recap-header{flex-wrap:nowrap!important}", css)
        self.assertIn(".recap-header .recap-date-filters{display:flex;align-items:center", css)
        self.assertIn("flex:0 0 160px;width:160px", css)
        self.assertIn("margin-left:auto!important", css)

    def test_interface_charge_actualise_les_trois_onglets_et_transmet_les_dates(self):
        javascript = (Path(settings.BASE_DIR) / "static/js/recapitulatif.js").read_text(encoding="utf-8")

        self.assertIn('new URLSearchParams({date_debut: selectedRange.start, date_fin: selectedRange.end})', javascript)
        self.assertIn('refreshButton?.addEventListener("click", () => selectRange(startInput.value, endInput.value))', javascript)
        self.assertIn('const selected = ["temps-travail", "paie", "centres", "totaux"].includes(tabName)', javascript)
        self.assertIn('button.addEventListener("click", () => openTab(button.dataset.recapTab))', javascript)
        self.assertLess(javascript.index("selectRange(startInput.value, endInput.value);"), javascript.index('openTab(new URLSearchParams'))
        self.assertIn('buildApiUrl("/recapitulatif/export-paie.pdf")', javascript)
        self.assertIn('buildApiUrl("/recapitulatif/export.xlsx")', javascript)

    def test_tableau_centres_separe_jours_details_contrat_et_reference_mensuelle(self):
        javascript = (Path(settings.BASE_DIR) / "static/js/recapitulatif.js").read_text(encoding="utf-8")
        css = (Path(settings.BASE_DIR) / "static/css/recapitulatif.css").read_text(encoding="utf-8")

        self.assertIn('<span class="paie-mobile-content"><span class="paie-cell-main">${result.jours_travailles}</span>${details}</span>', javascript)
        self.assertNotIn('${result.jours_travailles}${details', javascript)
        self.assertIn("segment.contrat_date_debut || segment.date_debut", javascript)
        self.assertIn("segment.contrat_date_fin || segment.date_fin", javascript)
        self.assertIn("Référence mensuelle</span>", javascript)
        self.assertIn('if (type === "apprentissage") return "Apprentissage";', javascript)
        self.assertNotIn('if (type === "apprentissage") return "Apprentissage / alternance";', javascript)
        self.assertIn("paie-pay-block paie-pay-block--reference", javascript)
        self.assertIn(".paie-pay-block--reference .paie-cell-main{color:var(--color-danger", css)
        self.assertIn("@media (max-width:799px)", css)
        self.assertIn(".recap-centres-table colgroup,.recap-centres-table thead{display:none}", css)
        self.assertIn("grid-template-columns:minmax(0,1fr) minmax(min-content,auto)", css)
        self.assertIn("td.complementary-column,.recap-centres-table tbody td.total-column", css)
        self.assertIn("justify-self:end;text-align:right", css)
        self.assertIn(".recap-centres-table{width:100%;table-layout:auto}", css)
        self.assertIn(".recap-centres-table .employee-col,.recap-centres-table .compact-col{width:1%}", css)
        self.assertNotIn("--recap-min-width", javascript)

    def test_tableau_centres_guide_les_preparations_incompletes(self):
        javascript = (Path(settings.BASE_DIR) / "static/js/recapitulatif.js").read_text(encoding="utf-8")
        css = (Path(settings.BASE_DIR) / "static/css/recapitulatif.css").read_text(encoding="utf-8")
        animateurs_js = (Path(settings.BASE_DIR) / "static/js/animateurs.js").read_text(encoding="utf-8")
        parametres_js = (Path(settings.BASE_DIR) / "static/js/parametres.js").read_text(encoding="utf-8")
        parametres_html = (Path(settings.BASE_DIR) / "templates/parametres.html").read_text(encoding="utf-8")

        self.assertIn('taux_contractuel_manquant: { label: "Taux contractuel manquant", section: "contrats" }', javascript)
        self.assertIn('salaire_mensuel_manquant: { label: "Salaire mensuel manquant", section: "contrats" }', javascript)
        self.assertIn('statut_manquant: { label: "Statut manquant", section: "historique-statut" }', javascript)
        self.assertIn('bareme_manquant: { label: "Barème CEE manquant", settings: true }', javascript)
        self.assertIn('paie-contract-segment--implicit', javascript)
        self.assertIn('CEE par défaut', javascript)
        self.assertIn('Contrat non renseigné', javascript)
        self.assertIn('payroll-preparation-row--incomplete', javascript)
        self.assertIn('payroll-preparation-row--verification', javascript)
        self.assertIn('.paie-non-calculable:focus-visible', css)
        self.assertIn('payroll-preparation-row--incomplete>th', css)

        self.assertIn('payroll-preparation-row--verification>th', css)
        self.assertIn('id="contrats"', animateurs_js)
        self.assertIn('id="historique-statut"', animateurs_js)
        self.assertIn('id="baremes-cee"', parametres_html)
        self.assertIn('window.location.hash === "#baremes-cee"', parametres_js)

        response = self.client.get(reverse("recapitulatif"))
        self.assertContains(response, 'data-employees-url="/employes/"')
        self.assertContains(response, 'data-settings-url="/parametres/"')

    def test_onglet_prime_rend_uniquement_les_eligibilites_du_backend(self):
        javascript = (Path(settings.BASE_DIR) / "static/js/recapitulatif.js").read_text(encoding="utf-8")
        css = (Path(settings.BASE_DIR) / "static/css/recapitulatif.css").read_text(encoding="utf-8")
        self.assertIn("primeEligibility = new Map", javascript)
        self.assertIn("contexte.primes", javascript)
        self.assertIn("prime.semaines_eligibles", javascript)
        self.assertIn("semaine.jours_eligibles.map", javascript)
        self.assertIn('amount.hidden = prime.type_montant === "fixe"', javascript)
        self.assertIn("Aucune nouvelle prime éligible sur cette période", javascript)
        self.assertIn("Primes à attribuer", javascript)
        self.assertIn("Historique des primes", javascript)
        self.assertIn("[Number(item.id), item]", javascript)
        self.assertNotIn("[Number(item.id), item.primes_eligibles || []]", javascript)
        self.assertEqual(javascript.count("Aucune nouvelle prime éligible sur cette période"), 1)
        self.assertIn("data-prime-id", javascript)
        self.assertIn("data-prime-entry-level", javascript)
        self.assertIn('niveau === "semaine"', javascript)
        self.assertIn('niveau === "jour"', javascript)
        self.assertNotIn("data-prime-type", javascript)
        self.assertIn("if (addAttribution.disabled) return", javascript)
        self.assertIn("setPrimeEditorBusy(editor, true)", javascript)
        self.assertIn("bodies = [{...baseBody, jours}]", javascript)
        self.assertIn("applyPrimeSynthesis(id, synthese)", javascript)
        self.assertIn('? "Enregistrement en attente…" : "Enregistrement…"', javascript)
        self.assertIn("showPrimeLocalError(form", javascript)
        self.assertIn("data-total-primes-animateur", javascript)
        self.assertIn("prime.jours_eligibles.length - joursDisponibles.length", javascript)
        self.assertIn("Aucun jour restant à attribuer", javascript)
        self.assertIn("prime.semaines_affichees", javascript)
        self.assertIn("synthese.contexte_prime", javascript)
        self.assertIn("Attributions en doublon sur cette période", javascript)
        self.assertIn("function primeSummaryValues(prime)", javascript)
        self.assertIn('detail: "aucune attribution"', javascript)
        self.assertIn("resume.montants_variables", javascript)
        self.assertIn("`${resume.quantite} j`", javascript)
        self.assertIn("updatePrimeEditorSummary(form.closest", javascript)
        self.assertIn('class="payroll-prime-summary-main"', javascript)
        self.assertIn('>(${escapeHtml(resume.detail)})</span>', javascript)
        self.assertIn(".payroll-prime-summary-main{display:flex", css)
        self.assertIn('mode === "jour"', javascript)
        self.assertIn("item.nombre_jours", javascript)
        self.assertIn("formatMoney(item.montant_unitaire)", javascript)
        self.assertIn("= ${formatMoney(item.montant_total)}", javascript)
        self.assertEqual(javascript.count("await loadRecap()"), 2)
        suppression = javascript.split('const deleteAttribution = event.target.closest("[data-delete-attribution]")', 1)[1].split('const editAttribution =', 1)[0]
        self.assertIn("enqueuePrimeDeletion", suppression)
        self.assertNotIn("setPrimeEditorBusy", suppression)
        self.assertNotIn("loadRecap", suppression)
        self.assertIn("const primeMutationQueues = new Map()", javascript)
        self.assertIn("const queuedPrimeMutations = new Set()", javascript)
        self.assertIn('key: `${id}:${typePrimeId}`', javascript)
        self.assertIn('"Suppression en attente…"', javascript)
        self.assertIn("async function processPrimeMutationQueue(key)", javascript)
        self.assertIn("function enqueuePrimeMutation(entry)", javascript)
        self.assertIn('key: `${id}:${prime.id}`', javascript)
        self.assertIn("const response = await entry.run()", javascript)
        self.assertIn("if (queuedPrimeMutations.has(entry.token)) return", javascript)
        self.assertIn('item?.setAttribute("aria-busy", "true")', javascript)
        self.assertIn("applyPrimeSynthesis(entry.animateurId, response.synthese)", javascript)
        self.assertIn("restoreQueuedDeletionStates()", javascript)
        self.assertIn("function optimisticAttribution(", javascript)
        self.assertIn('item.dataset.optimisticAttribution = ""', javascript)
        self.assertIn('"Enregistrement…"', javascript)
        self.assertIn('"Mise à jour…"', javascript)
        self.assertIn("optimistic?.rollback()", javascript)
        self.assertIn('item?.classList.add("is-pending-mutation")', javascript)
        self.assertIn("jours_eligibles: entry.joursEligibles", javascript)
        self.assertIn("jours_eligibles: prime.jours_eligibles || []", javascript)
        self.assertNotIn("contexteAnimateur.primes[index] = synthese.contexte_prime", javascript)
        self.assertIn("...contexteAnimateur.primes[index]", javascript)
        self.assertIn(".payroll-attribution.is-pending-mutation", css)

    def test_periode_paie_est_memorisee_et_restauree_avant_le_chargement(self):
        javascript = (Path(settings.BASE_DIR) / "static/js/recapitulatif.js").read_text(encoding="utf-8")
        self.assertIn('PAYROLL_RANGE_STORAGE_KEY = "animation-manager:payroll-date-range-v1"', javascript)
        self.assertIn("function isValidIsoDate(value)", javascript)
        self.assertIn("range.end < range.start", javascript)
        self.assertIn("localStorage.setItem(PAYROLL_RANGE_STORAGE_KEY", javascript)
        self.assertIn("const persistedRange = persistedPayrollRange();", javascript)
        self.assertIn("startInput.value = persistedRange?.start || formatIsoDate(firstDay)", javascript)
        self.assertLess(
            javascript.index("const persistedRange = persistedPayrollRange();"),
            javascript.rindex("selectRange(startInput.value, endInput.value);"),
        )

    def test_plage_calendaire_utilise_des_bornes_inclusives(self):
        self._affecter(self.julie, datetime.date(2026, 7, 6))
        self._affecter(self.julie, datetime.date(2026, 7, 20))

        response = self.client.get(
            reverse("api_recapitulatif") + "?date_debut=2026-07-01&date_fin=2026-07-31"
        )

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data["periode"]["debut"], "2026-07-01")
        self.assertEqual(data["periode"]["fin"], "2026-07-31")
        self.assertEqual(data["animateurs"][0]["jours_travailles"], 2)

    def test_mois_civil_calcule_le_premier_et_le_dernier_jour(self):
        response = self.client.get(reverse("api_recapitulatif") + "?mois=2028-02")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["periode"]["debut"], "2028-02-01")
        self.assertEqual(response.json()["periode"]["fin"], "2028-02-29")

    def test_plage_personnalisee_refuse_une_fin_anterieure(self):
        response = self.client.get(
            reverse("api_recapitulatif") + "?date_debut=2026-07-20&date_fin=2026-07-06"
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn("antérieure", response.json()["error"])

    def test_plage_calendaire_agrege_les_activites_de_plusieurs_semaines(self):
        for jour in (datetime.date(2026, 7, 6), datetime.date(2026, 7, 13), datetime.date(2026, 7, 27)):
            self._affecter(self.julie, jour)

        data = self.client.get(
            reverse("api_recapitulatif") + "?date_debut=2026-07-01&date_fin=2026-07-31"
        ).json()

        self.assertEqual(data["animateurs"][0]["jours_affectation"], 3)

    def test_plage_calendaire_ne_filtre_aucun_type_accueil(self):
        from animateurs.models import TypeAccueil

        types = [
            TypeAccueil.objects.get(code="vacances"),
            TypeAccueil.objects.get(code="periscolaire"),
            TypeAccueil.objects.get(code="sejours"),
        ]
        for index, type_accueil in enumerate(types):
            periode = creer_periode(
                debut=datetime.date(2026, 7, 6) + datetime.timedelta(days=7 * index),
                nom=f"Semaine {type_accueil.nom}",
            )
            periode.type_accueil = type_accueil
            periode.save(update_fields=["type_accueil"])
            self._affecter(self.julie, periode.debut)

        data = self.client.get(
            reverse("api_recapitulatif") + "?date_debut=2026-07-01&date_fin=2026-07-31"
        ).json()

        self.assertEqual(data["animateurs"][0]["jours_affectation"], 3)

    def test_prime_journaliere_est_limitee_et_propre_a_la_periode(self):
        self._affecter(self.julie, datetime.date(2026, 7, 6), duree=2)
        tarif_initial = Decimal(str(self.julie.paie_jour))
        url = reverse("api_prime_journaliere")

        response = self.client.put(
            url,
            data={"animateur_id": self.julie.id, "periode_ids": [self.periode.id], "montant": "5.00"},
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["total_jour_avec_prime"], "70.00")
        self.assertEqual(response.json()["total_paie_estime"], "140.00")
        self.assertEqual(
            PrimeJournalierePeriode.objects.get(animateur=self.julie, periode=self.periode).montant,
            5,
        )
        self.julie.refresh_from_db()
        self.assertEqual(self.julie.paie_jour, tarif_initial)

        invalide = self.client.put(
            url,
            data={"animateur_id": self.julie.id, "periode_ids": [self.periode.id], "montant": "7.01"},
            content_type="application/json",
        )
        self.assertEqual(invalide.status_code, 400)

    def test_api_recapitulatif_expose_la_preparation_de_paie(self):
        self._affecter(self.julie, datetime.date(2026, 7, 6))
        PrimeJournalierePeriode.objects.create(
            animateur=self.julie,
            periode=self.periode,
            montant="3",
        )

        data = self.client.get(
            reverse("api_recapitulatif") + f"?periode_ids={self.periode.id}"
        ).json()
        julie = next(item for item in data["animateurs"] if item["id"] == self.julie.id)

        self.assertEqual(julie["jours_travailles"], 1)
        self.assertEqual(julie["prime_jour"], "3.00")
        self.assertEqual(julie["total_jour_avec_prime"], "68.00")
        self.assertEqual(julie["total_paie_estime"], "68.00")

    def test_validation_groupee_applique_plusieurs_animateurs_et_semaines(self):
        seconde = creer_periode(debut=datetime.date(2026, 7, 13), nom="Été — Semaine 2")
        self._affecter(self.julie, datetime.date(2026, 7, 6))
        self._affecter(self.sam, datetime.date(2026, 7, 13))
        self.sam.paie_jour = Decimal("55.00")
        self.sam.save(update_fields=["paie_jour"])

        response = self.client.put(reverse("api_prime_journaliere"), data={
            "periode_ids": [self.periode.id, seconde.id],
            "primes": [
                {"animateur_id": self.julie.id, "montant": "3"},
                {"animateur_id": self.sam.id, "montant": "2.00"},
            ],
        }, content_type="application/json")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["semaines_modifiees"], 2)
        self.assertEqual(PrimeJournalierePeriode.objects.filter(
            animateur__in=[self.julie, self.sam], periode__in=[self.periode, seconde]
        ).count(), 4)

    def test_zero_groupe_supprime_uniquement_les_semaines_selectionnees(self):
        seconde = creer_periode(debut=datetime.date(2026, 7, 13), nom="Été — Semaine 2")
        PrimeJournalierePeriode.objects.create(animateur=self.julie, periode=self.periode, montant="2")
        PrimeJournalierePeriode.objects.create(animateur=self.julie, periode=seconde, montant="4")
        self._affecter(self.julie, datetime.date(2026, 7, 6))

        response = self.client.put(reverse("api_prime_journaliere"), data={
            "periode_ids": [self.periode.id],
            "primes": [{"animateur_id": self.julie.id, "montant": "0.00"}],
        }, content_type="application/json")

        self.assertEqual(response.status_code, 200)
        self.assertFalse(PrimeJournalierePeriode.objects.filter(animateur=self.julie, periode=self.periode).exists())
        self.assertTrue(PrimeJournalierePeriode.objects.filter(animateur=self.julie, periode=seconde, montant=4).exists())

    def test_validation_groupee_invalide_ne_modifie_rien(self):
        self._affecter(self.julie, datetime.date(2026, 7, 6))
        response = self.client.put(reverse("api_prime_journaliere"), data={
            "periode_ids": [self.periode.id],
            "primes": [
                {"animateur_id": self.julie.id, "montant": "3.00"},
                {"animateur_id": 999999, "montant": "2.00"},
            ],
        }, content_type="application/json")
        self.assertEqual(response.status_code, 400)
        self.assertFalse(PrimeJournalierePeriode.objects.exists())

    def test_primes_differentes_sont_detaillees_et_ponderees_par_semaine(self):
        seconde = creer_periode(debut=datetime.date(2026, 7, 13), nom="Été — Semaine 2")
        self._affecter(self.julie, datetime.date(2026, 7, 6), duree=5)
        self._affecter(self.julie, datetime.date(2026, 7, 13), duree=4)
        PrimeJournalierePeriode.objects.create(animateur=self.julie, periode=self.periode, montant="2")
        PrimeJournalierePeriode.objects.create(animateur=self.julie, periode=seconde, montant="5")

        data = self.client.get(reverse("api_recapitulatif") + f"?periode_ids={self.periode.id},{seconde.id}").json()
        julie = next(item for item in data["animateurs"] if item["id"] == self.julie.id)
        self.assertTrue(julie["prime_jour_variable"])
        self.assertEqual([item["prime_jour"] for item in julie["primes_detail"]], ["2.00", "5.00"])
        self.assertEqual(julie["paie_base"], "585.00")
        self.assertEqual(julie["montant_primes"], "30.00")
        self.assertEqual(julie["total_paie_estime"], "615.00")

    def test_page_prime_utilise_une_validation_explicite(self):
        response = self.client.get(reverse("recapitulatif"))
        self.assertContains(response, 'id="save-payroll-primes"')
        self.assertContains(response, "Valider les primes")
        self.assertContains(response, 'id="cancel-payroll-primes"')

    def test_suppression_individuelle_est_limitee_aux_semaines_selectionnees(self):
        seconde = creer_periode(debut=datetime.date(2026, 7, 13), nom="Été — Semaine 2")
        troisieme = creer_periode(debut=datetime.date(2026, 7, 20), nom="Été — Semaine 3")
        self._affecter(self.julie, datetime.date(2026, 7, 6))
        for periode, montant in ((self.periode, 2), (seconde, 4), (troisieme, 5)):
            PrimeJournalierePeriode.objects.create(animateur=self.julie, periode=periode, montant=montant)
        PrimeJournalierePeriode.objects.create(animateur=self.sam, periode=self.periode, montant=6)

        response = self.client.delete(reverse("api_prime_journaliere"), data={
            "animateur_id": self.julie.id, "periode_ids": [self.periode.id, seconde.id]
        }, content_type="application/json")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["deleted_count"], 2)
        self.assertFalse(PrimeJournalierePeriode.objects.filter(animateur=self.julie, periode__in=[self.periode, seconde]).exists())
        self.assertTrue(PrimeJournalierePeriode.objects.filter(animateur=self.julie, periode=troisieme, montant=5).exists())
        self.assertTrue(PrimeJournalierePeriode.objects.filter(animateur=self.sam, periode=self.periode, montant=6).exists())
        self.assertEqual(response.json()["animateur"]["montant_primes"], "0.00")

    def test_suppression_individuelle_tolere_une_semaine_sans_prime(self):
        seconde = creer_periode(debut=datetime.date(2026, 7, 13), nom="Été — Semaine 2")
        self._affecter(self.julie, datetime.date(2026, 7, 6))
        PrimeJournalierePeriode.objects.create(animateur=self.julie, periode=self.periode, montant=3)
        response = self.client.delete(reverse("api_prime_journaliere"), data={
            "animateur_id": self.julie.id, "periode_ids": [self.periode.id, seconde.id]
        }, content_type="application/json")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["deleted_count"], 1)

    def test_api_accepte_un_entier_et_refuse_toute_prime_decimale(self):
        self._affecter(self.julie, datetime.date(2026, 7, 6))
        url = reverse("api_prime_journaliere")
        payload = lambda montant: {"periode_ids": [self.periode.id], "primes": [{"animateur_id": self.julie.id, "montant": montant}]}
        self.assertEqual(self.client.put(url, data=payload(3), content_type="application/json").status_code, 200)
        for montant in ("2.50", "3.10", 3.5):
            response = self.client.put(url, data=payload(montant), content_type="application/json")
            self.assertEqual(response.status_code, 400)
            self.assertIn("euros entiers", response.json()["error"])
        self.assertEqual(PrimeJournalierePeriode.objects.get().montant, 3)

    def test_audit_primes_signale_sans_corriger_une_ancienne_decimale(self):
        prime = PrimeJournalierePeriode.objects.create(animateur=self.julie, periode=self.periode, montant="2.50")
        sortie = StringIO()
        call_command("audit_primes", stdout=sortie)
        self.assertIn("2,50 €", sortie.getvalue())
        prime.refresh_from_db()
        self.assertEqual(prime.montant, Decimal("2.50"))

    def test_prime_entiere_est_proratisee_sur_une_fraction_de_jour(self):
        self._affecter(self.julie, datetime.date(2026, 7, 6))
        preparation = ActiviteTravailComplementaire.objects.create(type="preparation", intitule="Demi-journée")
        preparation.periodes.add(self.periode)
        ParticipationTravailComplementaire.objects.create(
            activite=preparation, animateur=self.julie, nombre_jours="0.50"
        )
        PrimeJournalierePeriode.objects.create(animateur=self.julie, periode=self.periode, montant=3)
        data = self.client.get(reverse("api_recapitulatif") + f"?periode_ids={self.periode.id}").json()
        julie = next(item for item in data["animateurs"] if item["id"] == self.julie.id)
        self.assertEqual(julie["prime_jour"], "3.00")
        self.assertEqual(julie["montant_primes"], "4.50")
        self.assertEqual(julie["total_paie_estime"], "102.00")

    def test_interface_formate_la_prime_journaliere_sans_centimes(self):
        javascript = (Path(settings.BASE_DIR) / "static/js/recapitulatif.js").read_text(encoding="utf-8")
        self.assertIn('step="0.01"', javascript)
        self.assertIn("function formatDailyPrime", javascript)
        self.assertIn('`${number}\u00a0€`', javascript)
        self.assertIn("data-prime-id", javascript)
        self.assertIn("data-cancel-prime", javascript)
        self.assertIn("data-delete-prime", javascript)

    def test_prime_refusee_sans_jour_travaille_ou_tarif_journalier(self):
        url = reverse("api_prime_journaliere")
        sans_jour = self.client.put(
            url,
            data={"animateur_id": self.julie.id, "periode_ids": [self.periode.id], "montant": "2"},
            content_type="application/json",
        )
        self.assertEqual(sans_jour.status_code, 400)

        self._affecter(self.sam, datetime.date(2026, 7, 6))
        sans_tarif = self.client.put(
            url,
            data={"animateur_id": self.sam.id, "periode_ids": [self.periode.id], "montant": "2"},
            content_type="application/json",
        )
        self.assertEqual(sans_tarif.status_code, 400)

    def test_export_pdf_recapitulatif_paie(self):
        self._affecter(self.julie, datetime.date(2026, 7, 6), duree=2)

        response = self.client.get(
            reverse("export_recapitulatif_paie_pdf") + f"?periode_ids={self.periode.id}"
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/pdf")
        self.assertIn("recapitulatif_paie_20260706_20260710.pdf", response["Content-Disposition"])
        self.assertTrue(response.content.startswith(b"%PDF"))

    def test_pdf_utilise_base_cee_cp_primes_salaire_mensuel_et_alertes(self):
        def ligne(prenom, nom, **champs):
            return {
                "prenom": prenom, "nom": nom, "jours_affectation": 10,
                "jours_reunion": 0, "jours_preparation": 0, "jours_travailles": 10,
                "base_cee": "0.00", "indemnite_cp_cee": "0.00",
                "salaire_mensuel_reference": None, "base_mensuelle_reference": None,
                "montant_primes_preparees": "0.00", "total_prepare": "0.00",
                "etat_preparation": "pret", "paie_habituelle": False,
                "alertes_paie": [], **champs,
            }

        recap = {
            "animateurs": [
                ligne(
                    "Ange", "DUMONT", base_cee="600.00", indemnite_cp_cee="60.00",
                    montant_primes_preparees="80.00", total_prepare="740.00",
                    alertes_paie=[{
                        "message": "Contrat non renseigné — CEE appliqué par défaut",
                        "niveau": "information",
                    }, {
                        "message": "Statut historique incertain pour cette période.",
                        "niveau": "a_verifier",
                    }],
                ),
                ligne(
                    "Ambre", "BAIN", salaire_mensuel_reference="900.00",
                    base_mensuelle_reference="900.00", total_prepare="900.00",
                ),
                ligne(
                    "Paul", "PERMANENT", paie_habituelle=True,
                ),
            ],
            "total_jours": 30, "total_primes_preparees": "80.00",
            "total_prepare": "1640.00",
        }
        pdf = generer_recapitulatif_paie_pdf(
            recap, datetime.date(2026, 7, 1), datetime.date(2026, 7, 31)
        )
        lignes = lignes_recapitulatif_paie(recap)
        self.assertTrue(pdf.startswith(b"%PDF"))
        self.assertEqual(lignes[1][5:], [
            "600,00 €", "60,00 €", "—", "80,00 €", "740,00 €",
        ])
        self.assertEqual(lignes[2][7:], ["900,00 €", "0,00 €", "900,00 €"])
        self.assertEqual(lignes[3][-1], "Paie habituelle")
        self.assertNotIn("Paie par jour", lignes[0])
        self.assertEqual(alertes_recapitulatif_paie(recap), [{
            "animateur": "Ange DUMONT",
            "messages": [
                "Contrat non renseigné — CEE appliqué par défaut",
                "Statut historique incertain pour cette période.",
            ],
        }])

    def test_export_excel_recapitulatif(self):
        from openpyxl import load_workbook
        from io import BytesIO

        self._affecter(self.julie, datetime.date(2026, 7, 6), duree=2)
        response = self.client.get(
            reverse("export_recapitulatif_excel") + f"?periode_ids={self.periode.id}"
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("recapitulatif_20260706_20260710.xlsx", response["Content-Disposition"])
        classeur = load_workbook(BytesIO(response.content), data_only=True)
        self.assertEqual(classeur.sheetnames, ["Totaux paie", "Détail par centre", "Préparation contrats"])
        self.assertEqual(classeur["Totaux paie"]["A3"].value, "Julie Martin")
        self.assertEqual(classeur["Totaux paie"]["F2"].value, "Base CEE")
        self.assertEqual(classeur["Totaux paie"]["H2"].value, "Salaire mensuel de référence")
        self.assertNotIn("Paie par jour", [cell.value for cell in classeur["Totaux paie"][2]])

    def test_exports_acceptent_la_meme_plage_calendaire_inclusive(self):
        self._affecter(self.julie, datetime.date(2026, 7, 6), duree=2)
        query = "?date_debut=2026-07-01&date_fin=2026-07-31"

        pdf = self.client.get(reverse("export_recapitulatif_paie_pdf") + query)
        excel = self.client.get(reverse("export_recapitulatif_excel") + query)

        self.assertEqual(pdf.status_code, 200)
        self.assertEqual(excel.status_code, 200)
        self.assertIn("recapitulatif_paie_20260701_20260731.pdf", pdf["Content-Disposition"])
        self.assertIn("recapitulatif_20260701_20260731.xlsx", excel["Content-Disposition"])

    def test_tableau_pdf_agrege_exactement_les_totaux_ecran(self):
        seconde = creer_periode(debut=datetime.date(2026, 7, 13), nom="Été — Semaine 2")
        self._affecter(self.julie, datetime.date(2026, 7, 6), duree=5)
        self._affecter(self.julie, datetime.date(2026, 7, 13), duree=4)
        reunion = ActiviteTravailComplementaire.objects.create(
            type="reunion", intitule="Réunion", date=datetime.date(2026, 7, 7)
        )
        reunion.periodes.set([self.periode, seconde])
        ParticipationTravailComplementaire.objects.create(
            activite=reunion, animateur=self.julie, autoriser_double_comptage=True
        )
        PrimeJournalierePeriode.objects.create(animateur=self.julie, periode=self.periode, montant=3)
        PrimeJournalierePeriode.objects.create(animateur=self.julie, periode=seconde, montant=4)

        data = self.client.get(reverse("api_recapitulatif") + f"?periode_ids={self.periode.id},{seconde.id}").json()
        julie = next(item for item in data["animateurs"] if item["id"] == self.julie.id)
        lignes = lignes_recapitulatif_paie(data)

        self.assertEqual(len(lignes), 3)  # en-tête, un animateur, total général
        self.assertEqual(lignes[1], [
            "Julie Martin", "9", "1", "0", "10", "0,00 €",
            "0,00 €", "—", "34,00 €", "Incomplet",
        ])
        self.assertEqual(lignes[0], [
            "Animateur", "Affectations", "Réunions", "Télétravail / préparation",
            "Total jours", "Base CEE", "CP CEE", "Salaire mensuel de référence",
            "Primes", "Total préparé",
        ])
        self.assertEqual(julie["jours_travailles"], 10)
        self.assertEqual(julie["montant_primes"], "34.00")
        self.assertNotIn("0.125", " ".join(lignes[1]))
        self.assertFalse(any("Semaine" in cellule for ligne in lignes for cellule in ligne))

    def test_totaux_reutilisent_les_complements_temps_travail_dune_selection_englobante(self):
        seconde = creer_periode(debut=datetime.date(2026, 7, 13), nom="Été — Semaine 2")
        troisieme = creer_periode(debut=datetime.date(2026, 7, 20), nom="Été — Semaine 3")
        quatrieme = creer_periode(debut=datetime.date(2026, 7, 27), nom="Été — Semaine 4")
        aout = creer_periode(debut=datetime.date(2026, 8, 24), nom="Été — Août")
        ambre = Animateur.objects.create(prenom="Ambre", nom="BAIN")
        ange = Animateur.objects.create(prenom="Ange", nom="DUMONT")
        cassidy = Animateur.objects.create(prenom="Cassidy", nom="BOMBLE")
        jordan = Animateur.objects.create(prenom="Jordan", nom="SIMON")
        jours_ete = [
            periode.debut + datetime.timedelta(days=decalage)
            for periode in (self.periode, seconde, troisieme, quatrieme)
            for decalage in range(5)
        ]
        for animateur, nombre in ((ambre, 18), (ange, 10), (cassidy, 18), (jordan, 7)):
            for jour in jours_ete[:nombre]:
                self._affecter(animateur, jour)

        reunion = ActiviteTravailComplementaire.objects.create(
            type="reunion", intitule="Réunion été", date=datetime.date(2026, 7, 4)
        )
        selection_complete = [self.periode, seconde, troisieme, quatrieme, aout]
        reunion.periodes.set(selection_complete)
        for animateur in (ambre, ange, cassidy):
            ParticipationTravailComplementaire.objects.create(
                activite=reunion, animateur=animateur, nombre_jours="1.00"
            )
        preparation = ActiviteTravailComplementaire.objects.create(
            type="preparation", intitule="Prépa planning Ado"
        )
        preparation.periodes.set(selection_complete)
        ParticipationTravailComplementaire.objects.create(
            activite=preparation, animateur=jordan, nombre_jours="3.00"
        )
        ancienne_selection = ActiviteTravailComplementaire.objects.create(
            type="preparation", intitule="Ancienne sélection étroite"
        )
        ancienne_selection.periodes.set([seconde])
        ParticipationTravailComplementaire.objects.create(
            activite=ancienne_selection, animateur=jordan, nombre_jours="1.00"
        )

        data = self.client.get(reverse("api_recapitulatif"), {
            "date_debut": "2026-07-01", "date_fin": "2026-07-31",
        }).json()
        temps = self.client.get(reverse("api_temps_travail"), {
            "periode_ids": ",".join(str(item.id) for item in selection_complete),
        }).json()
        par_nom = {item["prenom"]: item for item in data["animateurs"]}
        synthese = {item["prenom"]: item for item in temps["synthese"]}
        attendus = {
            "Ambre": (18, 1, 0, 19), "Ange": (10, 1, 0, 11),
            "Cassidy": (18, 1, 0, 19), "Jordan": (7, 0, 3, 10),
        }
        for prenom, attendu in attendus.items():
            ligne = par_nom[prenom]
            self.assertEqual((
                ligne["jours_affectation"], ligne["jours_reunion"],
                ligne["jours_preparation"], ligne["jours_travailles"],
            ), attendu)
            self.assertEqual(ligne["jours_reunion"], synthese[prenom]["jours_reunion"])
            self.assertEqual(ligne["jours_preparation"], synthese[prenom]["jours_preparation"])
            self.assertEqual(ligne["jours_travailles"], synthese[prenom]["jours_total_recapitulatif"])

        lignes_export = {ligne[0]: ligne for ligne in lignes_recapitulatif_paie(data)[1:-1]}
        self.assertEqual(lignes_export["Jordan SIMON"][1:5], ["7", "0", "3", "10"])

    def test_reunion_n_est_pas_divisee_sur_les_semaines_vides(self):
        periodes = [self.periode] + [
            creer_periode(debut=datetime.date(2026, 7, 13) + datetime.timedelta(days=7 * index), nom=f"Été — Semaine {index + 2}")
            for index in range(7)
        ]
        self._affecter(self.julie, datetime.date(2026, 7, 6))
        reunion = ActiviteTravailComplementaire.objects.create(
            type="reunion", intitule="Réunion unique", date=datetime.date(2026, 7, 7)
        )
        reunion.periodes.set(periodes)
        ParticipationTravailComplementaire.objects.create(activite=reunion, animateur=self.julie)
        for periode in periodes:
            PrimeJournalierePeriode.objects.create(animateur=self.julie, periode=periode, montant=3)

        ids = ",".join(str(periode.id) for periode in periodes)
        julie = self.client.get(reverse("api_recapitulatif") + f"?periode_ids={ids}").json()["animateurs"][0]
        self.assertEqual(julie["jours_reunion"], 1)
        self.assertEqual(julie["montant_primes"], "6.00")
        self.assertEqual([item["jours"] for item in julie["primes_detail"]], [2, 0, 0, 0, 0, 0, 0, 0])

    def test_export_pdf_recapitulatif_refuse_une_periode_inconnue(self):
        response = self.client.get(reverse("export_recapitulatif_paie_pdf") + "?periode_ids=999999")

        self.assertEqual(response.status_code, 400)

    def test_api_refuse_une_selection_de_periode_inconnue(self):
        response = self.client.get(reverse("api_recapitulatif") + "?periode_ids=999999")
        self.assertEqual(response.status_code, 400)
