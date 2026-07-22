"""Génération et lecture des classeurs d'effectifs enfants."""

from __future__ import annotations

import io
import re
import unicodedata
from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path

import xlsxwriter
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import from_excel

from animateurs.models import Centre, EffectifEnfantsJour, Evenement, Groupe

TYPE_GABARIT = "animation_manager_effectifs"
VERSION_GABARIT = 1
NOM_FEUILLE_META = "_AM_META"
TAILLE_MAX_FICHIER = 10 * 1024 * 1024
NB_LIGNES_MAX = 5000

JOURS_FR = ("Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi", "Samedi", "Dimanche")


class ErreurExcel(ValueError):
    """Erreur métier lisible liée à un classeur importé."""


def normaliser_texte(value) -> str:
    texte = unicodedata.normalize("NFKD", str(value or ""))
    texte = "".join(char for char in texte if not unicodedata.combining(char))
    return re.sub(r"[^a-z0-9]+", " ", texte.lower()).strip()


def verifier_fichier_excel(fichier) -> None:
    nom = Path(getattr(fichier, "name", "")).name.lower()
    if not nom.endswith(".xlsx"):
        raise ErreurExcel("Seuls les fichiers Excel .xlsx sont acceptés.")
    taille = getattr(fichier, "size", None)
    if taille is not None and taille > TAILLE_MAX_FICHIER:
        raise ErreurExcel("Le fichier dépasse la taille maximale de 10 Mo.")


def _nom_feuille_unique(nom: str, utilises: set[str]) -> str:
    base = re.sub(r"[\\/*?:\[\]]", "-", nom).strip()[:31] or "Lieu"
    candidat = base
    suffixe = 2
    while candidat in utilises:
        fin = f" ({suffixe})"
        candidat = f"{base[:31-len(fin)]}{fin}"
        suffixe += 1
    utilises.add(candidat)
    return candidat


def _dates_gabarit(
    evenements: list[Evenement],
    debut: date,
    fin: date,
    dates_exclues: dict[int, set[date]],
) -> list[date]:
    jours = []
    courant = debut
    while courant <= fin:
        if any(evenement.est_ouvert_le(courant, dates_exclues[evenement.id]) for evenement in evenements):
            jours.append(courant)
        courant += timedelta(days=1)
    return jours


def generer_gabarit_excel(centre_ids: list[int], debut: date, fin: date) -> bytes:
    if not centre_ids:
        raise ErreurExcel("Sélectionnez au moins un lieu.")
    if fin < debut:
        raise ErreurExcel("La date de fin doit être postérieure à la date de début.")
    if (fin - debut).days > 366:
        raise ErreurExcel("Le gabarit ne peut pas couvrir plus d'un an.")

    evenements = (
        Evenement.objects.filter(centre_id__in=centre_ids)
        .select_related("centre", "groupe")
        .prefetch_related("periodes_scolaires", "dates_exclues")
        .order_by("centre__ordre", "centre__nom", "ordre", "nom")
    )
    par_centre: dict[int, list[Evenement]] = defaultdict(list)
    for evenement in evenements:
        par_centre[evenement.centre_id].append(evenement)

    centres = list(Centre.objects.filter(id__in=centre_ids).order_by("ordre", "nom"))
    manquants = [centre.nom for centre in centres if not par_centre.get(centre.id)]
    if manquants:
        raise ErreurExcel(f"Aucun groupe n'est configuré pour : {', '.join(manquants)}.")

    effectifs = {
        (item.evenement_id, item.date): item.nombre
        for item in EffectifEnfantsJour.objects.filter(
            evenement__centre_id__in=centre_ids,
            date__gte=debut,
            date__lte=fin,
        )
    }

    sortie = io.BytesIO()
    classeur = xlsxwriter.Workbook(sortie, {"in_memory": True})
    classeur.set_properties({
        "title": "Gabarit d'effectifs enfants",
        "subject": "Animation Manager",
        "comments": "Généré automatiquement par Animation Manager",
    })
    fmt_titre = classeur.add_format({"bold": True, "font_size": 16, "font_color": "#4F3AA8"})
    fmt_info = classeur.add_format({"font_color": "#667085", "italic": True})
    fmt_entete = classeur.add_format({
        "bold": True, "bg_color": "#EEEAFB", "font_color": "#403073",
        "border": 1, "border_color": "#D6D0EB", "align": "center", "valign": "vcenter",
    })
    fmt_date = classeur.add_format({"num_format": "dd/mm/yyyy", "border": 1, "border_color": "#E4E7EC"})
    fmt_jour = classeur.add_format({"border": 1, "border_color": "#E4E7EC", "font_color": "#667085"})
    fmt_nombre = classeur.add_format({"border": 1, "border_color": "#E4E7EC", "align": "center"})
    fmt_ferme = classeur.add_format({"border": 1, "border_color": "#E4E7EC", "bg_color": "#F2F4F7"})
    fmt_note = classeur.add_format({"font_color": "#667085", "font_size": 9})

    meta = classeur.add_worksheet(NOM_FEUILLE_META)
    meta.write_row(0, 0, ["type", TYPE_GABARIT])
    meta.write_row(1, 0, ["version", VERSION_GABARIT])
    meta.write_row(2, 0, ["debut", debut.isoformat()])
    meta.write_row(3, 0, ["fin", fin.isoformat()])
    meta.write_row(5, 0, ["sheet", "centre_id", "column", "evenement_id", "groupe_id"])
    ligne_meta = 6
    utilises: set[str] = set()

    for centre in centres:
        groupes = par_centre[centre.id]
        nom_feuille = _nom_feuille_unique(centre.nom, utilises)
        feuille = classeur.add_worksheet(nom_feuille)
        feuille.freeze_panes(4, 2)
        feuille.hide_gridlines(2)
        feuille.set_tab_color(centre.couleur)
        feuille.merge_range(0, 0, 0, max(2, len(groupes) + 1), f"Effectifs enfants — {centre.nom}", fmt_titre)
        feuille.write(1, 0, f"Du {debut:%d/%m/%Y} au {fin:%d/%m/%Y}", fmt_info)
        feuille.write(2, 0, "Remplissez uniquement les colonnes des groupes. Une cellule vide ne modifiera rien lors de l'import.", fmt_note)
        feuille.write_row(3, 0, ["Date", "Jour", *[groupe.nom for groupe in groupes]], fmt_entete)
        feuille.set_column(0, 0, 13)
        feuille.set_column(1, 1, 13)
        feuille.set_column(2, len(groupes) + 1, 18)
        feuille.autofilter(3, 0, 3, len(groupes) + 1)

        dates_exclues_par_evenement = {
            evenement.id: {fermeture.date for fermeture in evenement.dates_exclues.all()}
            for evenement in groupes
        }
        jours = _dates_gabarit(groupes, debut, fin, dates_exclues_par_evenement)
        for index, jour in enumerate(jours, start=4):
            feuille.write_datetime(index, 0, datetime.combine(jour, datetime.min.time()), fmt_date)
            feuille.write(index, 1, JOURS_FR[jour.weekday()], fmt_jour)
            for offset, evenement in enumerate(groupes, start=2):
                ouvert = evenement.est_ouvert_le(jour, dates_exclues_par_evenement[evenement.id])
                if not ouvert:
                    feuille.write_blank(index, offset, None, fmt_ferme)
                    continue
                valeur = effectifs.get((evenement.id, jour))
                if valeur is None:
                    feuille.write_blank(index, offset, None, fmt_nombre)
                else:
                    feuille.write_number(index, offset, valeur, fmt_nombre)
                feuille.data_validation(index, offset, index, offset, {
                    "validate": "integer", "criteria": "between", "minimum": 0, "maximum": 999,
                    "input_title": "Effectif", "input_message": "Saisissez un nombre entre 0 et 999.",
                    "error_title": "Valeur incorrecte", "error_message": "Saisissez un nombre entier entre 0 et 999.",
                })

        for offset, evenement in enumerate(groupes, start=2):
            meta.write_row(ligne_meta, 0, [nom_feuille, centre.id, offset + 1, evenement.id, evenement.groupe_id])
            ligne_meta += 1

    meta.hide()
    classeur.close()
    return sortie.getvalue()


def _charger_classeur(fichier):
    verifier_fichier_excel(fichier)
    try:
        fichier.seek(0)
        return load_workbook(fichier, read_only=True, data_only=True)
    except Exception as exc:  # openpyxl lève plusieurs types selon la corruption.
        raise ErreurExcel("Le fichier Excel est illisible ou endommagé.") from exc


def _ligne_entete(feuille) -> int | None:
    for numero, ligne in enumerate(feuille.iter_rows(min_row=1, max_row=min(20, feuille.max_row), values_only=True), start=1):
        non_vides = [value for value in ligne if value not in (None, "")]
        if len(non_vides) >= 2:
            return numero
    return None


def _metadata_officielles(classeur) -> dict | None:
    if NOM_FEUILLE_META not in classeur.sheetnames:
        return None
    feuille = classeur[NOM_FEUILLE_META]
    if feuille.cell(1, 1).value != "type" or feuille.cell(1, 2).value != TYPE_GABARIT:
        return None
    if int(feuille.cell(2, 2).value or 0) != VERSION_GABARIT:
        raise ErreurExcel("Cette version du gabarit n'est pas compatible avec l'application.")
    colonnes = []
    for ligne in feuille.iter_rows(min_row=7, values_only=True):
        if not ligne or not ligne[0]:
            continue
        colonnes.append({
            "sheet": str(ligne[0]),
            "centre_id": int(ligne[1]),
            "column": int(ligne[2]),
            "evenement_id": int(ligne[3]),
            "groupe_id": int(ligne[4]),
        })
    return {"columns": colonnes}


def analyser_classeur(fichier) -> dict:
    classeur = _charger_classeur(fichier)
    meta = _metadata_officielles(classeur)
    feuilles = []
    for nom in classeur.sheetnames:
        if nom == NOM_FEUILLE_META:
            continue
        feuille = classeur[nom]
        entete = 4 if meta else _ligne_entete(feuille)
        if not entete:
            feuilles.append({"name": nom, "header_row": None, "columns": [], "samples": [], "values": {}})
            continue
        valeurs_entete = next(feuille.iter_rows(min_row=entete, max_row=entete, values_only=True))
        colonnes = []
        for index, value in enumerate(valeurs_entete, start=1):
            if value in (None, ""):
                continue
            colonnes.append({"letter": get_column_letter(index), "index": index, "label": str(value).strip()})
        samples = []
        valeurs_uniques = {col["letter"]: [] for col in colonnes}
        vus = {col["letter"]: set() for col in colonnes}
        for numero, ligne in enumerate(
            feuille.iter_rows(min_row=entete + 1, max_row=min(feuille.max_row, entete + 500), values_only=True),
            start=entete + 1,
        ):
            if not any(value not in (None, "") for value in ligne):
                continue
            sample = {"row": numero}
            for col in colonnes:
                value = ligne[col["index"] - 1] if col["index"] - 1 < len(ligne) else None
                affichage = _afficher_valeur(value)
                sample[col["letter"]] = affichage
                cle = str(affichage)
                if affichage not in (None, "") and cle not in vus[col["letter"]] and len(valeurs_uniques[col["letter"]]) < 50:
                    vus[col["letter"]].add(cle)
                    valeurs_uniques[col["letter"]].append(affichage)
            if len(samples) < 5:
                samples.append(sample)
        feuilles.append({
            "name": nom,
            "header_row": entete,
            "columns": colonnes,
            "samples": samples,
            "values": valeurs_uniques,
        })
    return {"official": bool(meta), "sheets": feuilles}


def _afficher_valeur(value):
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return value


def _col_index(letter: str) -> int:
    lettre = str(letter or "").strip().upper()
    if not re.fullmatch(r"[A-Z]{1,3}", lettre):
        raise ErreurExcel(f"Colonne Excel invalide : {letter}.")
    resultat = 0
    for char in lettre:
        resultat = resultat * 26 + ord(char) - 64
    return resultat


def _valeur_ligne(ligne, letter: str):
    index = _col_index(letter) - 1
    return ligne[index] if index < len(ligne) else None


def _parse_date(value, epoch=None) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, int | float) and not isinstance(value, bool):
        try:
            return from_excel(value, epoch=epoch).date()
        except Exception:
            return None
    texte = str(value).strip()
    for format_ in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y", "%d/%m/%y", "%d-%m-%y"):
        try:
            return datetime.strptime(texte, format_).date()
        except ValueError:
            continue
    return None


def _parse_nombre(value) -> int | None:
    if value in (None, ""):
        return None
    if isinstance(value, bool):
        raise ErreurExcel("Une valeur vrai/faux ne peut pas être utilisée comme effectif.")
    if isinstance(value, int):
        nombre = value
    elif isinstance(value, float):
        if not value.is_integer():
            raise ErreurExcel("Les effectifs doivent être des nombres entiers.")
        nombre = int(value)
    else:
        texte = str(value).strip().replace(" ", "")
        if not re.fullmatch(r"\d+", texte):
            raise ErreurExcel(f"« {value} » n'est pas un effectif valide.")
        nombre = int(texte)
    if not 0 <= nombre <= 999:
        raise ErreurExcel("Les effectifs doivent être compris entre 0 et 999.")
    return nombre


def _correspondance_id(value, mappings: dict, automatiques: dict[str, int]) -> int | None:
    texte = str(_afficher_valeur(value) or "").strip()
    if texte in mappings:
        try:
            return int(mappings[texte])
        except (TypeError, ValueError):
            return None
    return automatiques.get(normaliser_texte(texte))


def _ajouter_ligne(resultats, erreurs, donnees, source, cle_doublons):
    centre_id, groupe_id, jour, nombre = donnees
    if not all((centre_id, groupe_id, jour is not None, nombre is not None)):
        return
    cle = (centre_id, groupe_id, jour)
    existante = cle_doublons.get(cle)
    if existante is not None:
        if existante != nombre:
            erreurs.append({**source, "message": "Cette date et ce groupe apparaissent plusieurs fois avec des effectifs différents."})
        return
    cle_doublons[cle] = nombre
    resultats.append({"centre_id": centre_id, "groupe_id": groupe_id, "date": jour, "nombre": nombre, **source})


def _lignes_officielles(classeur, meta) -> tuple[list[dict], list[dict]]:
    resultats, erreurs, doublons = [], [], {}
    groupes_par_feuille = defaultdict(list)
    for config in meta["columns"]:
        groupes_par_feuille[config["sheet"]].append(config)
    evenements = {item.id: item for item in Evenement.objects.select_related("centre", "groupe").filter(
        id__in=[item["evenement_id"] for item in meta["columns"]]
    )}
    for nom_feuille, colonnes in groupes_par_feuille.items():
        if nom_feuille not in classeur.sheetnames:
            erreurs.append({"sheet": nom_feuille, "row": None, "message": "La feuille attendue est absente du gabarit."})
            continue
        feuille = classeur[nom_feuille]
        for numero, ligne in enumerate(feuille.iter_rows(min_row=5, values_only=True), start=5):
            jour = _parse_date(ligne[0] if ligne else None, classeur.epoch)
            if not jour:
                if any(value not in (None, "") for value in ligne[2:]):
                    erreurs.append({"sheet": nom_feuille, "row": numero, "message": "La date de cette ligne est invalide."})
                continue
            for config in colonnes:
                evenement = evenements.get(config["evenement_id"])
                if not evenement or evenement.centre_id != config["centre_id"] or evenement.groupe_id != config["groupe_id"]:
                    erreurs.append({"sheet": nom_feuille, "row": numero, "message": "Le groupe lié au gabarit n'existe plus."})
                    continue
                value = ligne[config["column"] - 1] if config["column"] - 1 < len(ligne) else None
                try:
                    nombre = _parse_nombre(value)
                except ErreurExcel as exc:
                    erreurs.append({"sheet": nom_feuille, "row": numero, "message": f"{evenement.nom} : {exc}"})
                    continue
                if nombre is None:
                    continue
                _ajouter_ligne(resultats, erreurs, (evenement.centre_id, evenement.groupe_id, jour, nombre), {
                    "sheet": nom_feuille, "row": numero,
                }, doublons)
    return resultats, erreurs


def _lignes_externes(classeur, configuration: dict) -> tuple[list[dict], list[dict]]:
    resultats, erreurs, doublons = [], [], {}
    centres = list(Centre.objects.all())
    groupes = list(Groupe.objects.all())
    centres_auto = {}
    for centre in centres:
        centres_auto[normaliser_texte(centre.nom)] = centre.id
        centres_auto[normaliser_texte(centre.code)] = centre.id
    groupes_auto = {normaliser_texte(groupe.nom): groupe.id for groupe in groupes}

    for config in configuration.get("sheets", []):
        if not config.get("enabled", True):
            continue
        nom = str(config.get("name", ""))
        if nom not in classeur.sheetnames or nom == NOM_FEUILLE_META:
            erreurs.append({"sheet": nom, "row": None, "message": "Feuille introuvable."})
            continue
        feuille = classeur[nom]
        try:
            entete = int(config.get("header_row") or 1)
            date_col = config["date_column"]
            layout = config.get("layout", "wide")
        except (KeyError, TypeError, ValueError):
            erreurs.append({"sheet": nom, "row": None, "message": "La configuration de la feuille est incomplète."})
            continue

        centre_cfg = config.get("centre") or {}
        centre_mode = centre_cfg.get("mode", "fixed")
        fixed_centre = None
        if centre_mode == "fixed":
            try:
                fixed_centre = int(centre_cfg.get("centre_id"))
            except (TypeError, ValueError):
                erreurs.append({"sheet": nom, "row": None, "message": "Choisissez le lieu correspondant à cette feuille."})
                continue
        centre_mappings = {str(k): v for k, v in (centre_cfg.get("values") or {}).items()}

        for numero, ligne in enumerate(feuille.iter_rows(min_row=entete + 1, values_only=True), start=entete + 1):
            if not any(value not in (None, "") for value in ligne):
                continue
            jour = _parse_date(_valeur_ligne(ligne, date_col), classeur.epoch)
            if not jour:
                erreurs.append({"sheet": nom, "row": numero, "message": "Date invalide ou non reconnue."})
                continue
            if centre_mode == "fixed":
                centre_id = fixed_centre
            else:
                centre_value = _valeur_ligne(ligne, centre_cfg.get("column"))
                centre_id = _correspondance_id(centre_value, centre_mappings, centres_auto)
                if not centre_id:
                    erreurs.append({"sheet": nom, "row": numero, "message": f"Lieu non reconnu : « {_afficher_valeur(centre_value)} »."})
                    continue

            if layout == "long":
                groupe_cfg = config.get("group") or {}
                groupe_value = _valeur_ligne(ligne, groupe_cfg.get("column"))
                groupe_id = _correspondance_id(
                    groupe_value,
                    {str(k): v for k, v in (groupe_cfg.get("values") or {}).items()},
                    groupes_auto,
                )
                if not groupe_id:
                    erreurs.append({"sheet": nom, "row": numero, "message": f"Groupe non reconnu : « {_afficher_valeur(groupe_value)} »."})
                    continue
                try:
                    nombre = _parse_nombre(_valeur_ligne(ligne, config.get("effectif_column")))
                except ErreurExcel as exc:
                    erreurs.append({"sheet": nom, "row": numero, "message": str(exc)})
                    continue
                if nombre is not None:
                    _ajouter_ligne(resultats, erreurs, (centre_id, groupe_id, jour, nombre), {"sheet": nom, "row": numero}, doublons)
            else:
                for colonne, groupe_id_brut in (config.get("group_columns") or {}).items():
                    try:
                        groupe_id = int(groupe_id_brut)
                    except (TypeError, ValueError):
                        continue
                    try:
                        nombre = _parse_nombre(_valeur_ligne(ligne, colonne))
                    except ErreurExcel as exc:
                        erreurs.append({"sheet": nom, "row": numero, "message": f"Colonne {colonne} : {exc}"})
                        continue
                    if nombre is not None:
                        _ajouter_ligne(resultats, erreurs, (centre_id, groupe_id, jour, nombre), {"sheet": nom, "row": numero}, doublons)

            if len(resultats) > NB_LIGNES_MAX:
                raise ErreurExcel(f"L'import est limité à {NB_LIGNES_MAX} effectifs.")
    return resultats, erreurs


def previsualiser_classeur(fichier, configuration: dict | None = None) -> dict:
    classeur = _charger_classeur(fichier)
    meta = _metadata_officielles(classeur)
    if meta:
        brutes, erreurs = _lignes_officielles(classeur, meta)
        officiel = True
    else:
        if not configuration:
            raise ErreurExcel("Indiquez la correspondance des colonnes du fichier.")
        brutes, erreurs = _lignes_externes(classeur, configuration)
        officiel = False

    couples = {(item["centre_id"], item["groupe_id"]) for item in brutes}
    evenements = {
        (item.centre_id, item.groupe_id): item
        for item in Evenement.objects.select_related("centre", "groupe").filter(
            centre_id__in={centre_id for centre_id, _ in couples},
            groupe_id__in={groupe_id for _, groupe_id in couples},
        )
    }
    dates = [item["date"] for item in brutes]
    existants = {}
    if dates:
        ids_evenements = [item.id for item in evenements.values()]
        existants = {
            (item.evenement_id, item.date): item.nombre
            for item in EffectifEnfantsJour.objects.filter(
                evenement_id__in=ids_evenements,
                date__gte=min(dates),
                date__lte=max(dates),
            )
        }

    lignes = []
    for brute in brutes:
        evenement = evenements.get((brute["centre_id"], brute["groupe_id"]))
        if not evenement:
            erreurs.append({
                "sheet": brute["sheet"], "row": brute["row"],
                "message": "Ce groupe n'est pas configuré dans le lieu indiqué.",
            })
            continue
        actuel = existants.get((evenement.id, brute["date"]), 0)
        lignes.append({
            "evenement_id": evenement.id,
            "centre_id": evenement.centre_id,
            "centre_nom": evenement.centre.nom,
            "groupe_id": evenement.groupe_id,
            "groupe_nom": evenement.groupe.nom,
            "date": brute["date"].isoformat(),
            "actuel": actuel,
            "importe": brute["nombre"],
            "change": actuel != brute["nombre"],
            "sheet": brute["sheet"],
            "row": brute["row"],
        })
    lignes.sort(key=lambda item: (item["date"], item["centre_nom"], item["groupe_nom"]))
    return {"official": officiel, "rows": lignes, "errors": erreurs}
