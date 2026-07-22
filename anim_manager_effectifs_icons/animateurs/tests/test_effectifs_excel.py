import io
import json
from datetime import date

from django.urls import reverse
from openpyxl import Workbook, load_workbook

from animateurs.models import Centre, EffectifEnfantsJour, Evenement, ProfilImportEffectifs
from animateurs.tests.base import ConnexionTestCase


class EffectifsExcelTests(ConnexionTestCase):
    def setUp(self):
        self.lp = Centre.objects.create(nom="La Pacaudière", code="LP", ordre=0)
        self.sf = Centre.objects.create(nom="Saint-Forgeux-les-Pins", code="SF", ordre=1)
        self.lp_mat = Evenement.objects.create(centre=self.lp, nom="Maternels", permanent=True, jours_ouverts=[0, 1, 2, 3, 4])
        self.lp_elem = Evenement.objects.create(centre=self.lp, nom="Élémentaires", permanent=True, jours_ouverts=[0, 1, 2, 3, 4])
        self.sf_mat = Evenement.objects.create(centre=self.sf, nom="Maternels", permanent=True, jours_ouverts=[0, 1, 2, 3, 4])
        self.sf_elem = Evenement.objects.create(centre=self.sf, nom="Élémentaires", permanent=True, jours_ouverts=[0, 1, 2, 3, 4])

    def _xlsx_file(self, workbook, name="effectifs.xlsx"):
        stream = io.BytesIO()
        workbook.save(stream)
        stream.seek(0)
        stream.name = name
        return stream

    def test_gabarit_multi_lieux_et_previsualisation_officielle(self):
        EffectifEnfantsJour.objects.create(
            evenement=self.lp_mat,
            date=date(2026, 7, 6),
            nombre=12,
            enfants_par_animateur=8,
        )
        response = self.client.get(reverse("api_effectifs_excel_gabarit"), {
            "debut": "2026-07-06",
            "fin": "2026-07-10",
            "centre": [self.lp.id, self.sf.id],
        })
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        workbook = load_workbook(io.BytesIO(response.content))
        self.assertIn("La Pacaudière", workbook.sheetnames)
        self.assertIn("Saint-Forgeux-les-Pins", workbook.sheetnames)
        self.assertEqual(workbook["La Pacaudière"]["C5"].value, 12)
        workbook["La Pacaudière"]["C5"] = 18
        fichier = self._xlsx_file(workbook)

        analyse = self.client.post(reverse("api_effectifs_excel_analyser"), {"fichier": fichier})
        self.assertEqual(analyse.status_code, 200)
        self.assertTrue(analyse.json()["official"])

        fichier.seek(0)
        preview = self.client.post(reverse("api_effectifs_excel_previsualiser"), {"fichier": fichier})
        self.assertEqual(preview.status_code, 200)
        ligne = next(item for item in preview.json()["rows"] if item["evenement_id"] == self.lp_mat.id and item["date"] == "2026-07-06")
        self.assertEqual(ligne["actuel"], 12)
        self.assertEqual(ligne["importe"], 18)
        self.assertTrue(ligne["change"])

    def test_import_externe_large_multi_lieux(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Export logiciel"
        sheet.append(["Date", "Site", "3-5 ans", "6-10 ans"])
        sheet.append([date(2026, 7, 6), "LP", 14, 25])
        sheet.append([date(2026, 7, 6), "SF", 11, 20])
        fichier = self._xlsx_file(workbook)

        configuration = {
            "sheets": [{
                "name": "Export logiciel",
                "enabled": True,
                "header_row": 1,
                "date_column": "A",
                "centre": {"mode": "column", "column": "B", "values": {"LP": self.lp.id, "SF": self.sf.id}},
                "layout": "wide",
                "group_columns": {"C": self.lp_mat.groupe_id, "D": self.lp_elem.groupe_id},
                "group": {"column": "", "values": {}},
                "effectif_column": "",
            }]
        }
        response = self.client.post(reverse("api_effectifs_excel_previsualiser"), {
            "fichier": fichier,
            "configuration": json.dumps(configuration),
        })
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["errors"], [])
        self.assertEqual(len(payload["rows"]), 4)
        self.assertEqual({(item["centre_nom"], item["groupe_nom"], item["importe"]) for item in payload["rows"]}, {
            ("La Pacaudière", "Maternels", 14),
            ("La Pacaudière", "Élémentaires", 25),
            ("Saint-Forgeux-les-Pins", "Maternels", 11),
            ("Saint-Forgeux-les-Pins", "Élémentaires", 20),
        })

    def test_import_externe_long_et_validation(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "LP"
        sheet.append(["Journée", "Tranche", "Nombre"])
        sheet.append(["06/07/2026", "MAT", 16])
        sheet.append(["06/07/2026", "ELEM", 29])
        fichier = self._xlsx_file(workbook)
        configuration = {
            "sheets": [{
                "name": "LP",
                "enabled": True,
                "header_row": 1,
                "date_column": "A",
                "centre": {"mode": "fixed", "centre_id": self.lp.id},
                "layout": "long",
                "group_columns": {},
                "group": {"column": "B", "values": {"MAT": self.lp_mat.groupe_id, "ELEM": self.lp_elem.groupe_id}},
                "effectif_column": "C",
            }]
        }
        response = self.client.post(reverse("api_effectifs_excel_previsualiser"), {
            "fichier": fichier,
            "configuration": json.dumps(configuration),
        })
        self.assertEqual(response.status_code, 200)
        rows = response.json()["rows"]
        import_response = self.client.post(
            reverse("api_effectifs_excel_importer"),
            data=json.dumps({"rows": rows}),
            content_type="application/json",
        )
        self.assertEqual(import_response.status_code, 200)
        self.assertEqual(
            import_response.json()["periodes"],
            [{"debut": "2026-07-06", "fin": "2026-07-13"}],
        )
        self.assertEqual(EffectifEnfantsJour.objects.get(evenement=self.lp_mat, date=date(2026, 7, 6)).nombre, 16)
        self.assertEqual(EffectifEnfantsJour.objects.get(evenement=self.lp_elem, date=date(2026, 7, 6)).nombre, 29)

    def test_import_retourne_toutes_les_semaines_touchees(self):
        response = self.client.post(
            reverse("api_effectifs_excel_importer"),
            data=json.dumps({
                "rows": [
                    {"evenement_id": self.lp_mat.id, "date": "2026-07-06", "importe": 12},
                    {"evenement_id": self.lp_mat.id, "date": "2026-07-15", "importe": 18},
                    {"evenement_id": self.sf_elem.id, "date": "2026-07-27", "importe": 24},
                ]
            }),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["periodes"], [
            {"debut": "2026-07-06", "fin": "2026-07-13"},
            {"debut": "2026-07-13", "fin": "2026-07-20"},
            {"debut": "2026-07-27", "fin": "2026-08-03"},
        ])

    def test_zero_supprime_effectif_sans_perdre_un_ratio_exceptionnel(self):
        ligne = EffectifEnfantsJour.objects.create(
            evenement=self.lp_mat,
            date=date(2026, 7, 6),
            nombre=12,
            enfants_par_animateur=6,
            ratio_encadrement_exceptionnel=6,
        )
        response = self.client.post(
            reverse("api_effectifs_excel_importer"),
            data=json.dumps({"rows": [{"evenement_id": self.lp_mat.id, "date": "2026-07-06", "importe": 0}]}),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        ligne.refresh_from_db()
        self.assertEqual(ligne.nombre, 0)
        self.assertEqual(ligne.ratio_encadrement_exceptionnel, 6)

    def test_profils_sont_propres_a_utilisateur(self):
        configuration = {"sheets": [{"name": "Export"}]}
        response = self.client.post(
            reverse("api_profils_import_effectifs"),
            data=json.dumps({"nom": "Export Millibase", "configuration": configuration}),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 201)
        profil = ProfilImportEffectifs.objects.get()
        self.assertEqual(profil.utilisateur, self.compte_maitre)
        self.assertEqual(self.client.get(reverse("api_profils_import_effectifs")).json()[0]["nom"], "Export Millibase")
        delete = self.client.delete(reverse("api_profil_import_effectifs_detail", args=[profil.id]))
        self.assertEqual(delete.status_code, 200)
        self.assertFalse(ProfilImportEffectifs.objects.exists())

    def test_page_planning_contient_assistant_excel(self):
        response = self.client.get(reverse("planning"))
        self.assertContains(response, 'id="btn-effectifs-excel"')
        self.assertContains(response, 'id="modal-effectifs-excel"')
        self.assertContains(response, "js/effectifs-excel.js")

    def test_analyse_externe_detecte_colonnes_et_valeurs(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Export"
        sheet.append(["Date", "Lieu", "Maternels", "Élémentaires"])
        sheet.append([date(2026, 7, 6), "LP", 12, 24])
        fichier = self._xlsx_file(workbook)
        response = self.client.post(reverse("api_effectifs_excel_analyser"), {"fichier": fichier})
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertFalse(payload["official"])
        self.assertEqual(payload["sheets"][0]["header_row"], 1)
        self.assertEqual(payload["sheets"][0]["values"]["B"], ["LP"])
        self.assertEqual({item["code"] for item in payload["centres"]}, {"LP", "SF"})

    def test_rejette_les_anciens_formats_excel(self):
        fichier = io.BytesIO(b"pas un classeur")
        fichier.name = "effectifs.xls"
        response = self.client.post(reverse("api_effectifs_excel_analyser"), {"fichier": fichier})
        self.assertEqual(response.status_code, 400)
        self.assertIn(".xlsx", response.json()["error"])
